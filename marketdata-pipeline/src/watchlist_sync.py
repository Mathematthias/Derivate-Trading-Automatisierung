"""
watchlist_sync.py — Synchronisiert die Watchlist aus dem Trading-Journal
(Excel) in den Watchlist-Block des STATE-Doc.

Workflow:
1. Sucht das jüngste Trading_Journal_*.xlsx im Workspace-Shared-Drive
2. Lädt es runter, liest das "Watchlist"-Sheet
3. Konvertiert Excel-Zeilen in das STATE-Markdown-Tabellenformat
4. Liest STATE-Doc, ersetzt nur die Watchlist-Sektion
5. Schreibt STATE-Doc zurück via Drive API files().update()

Aufruf in GitHub Action (vor marketdata_sync.py):
    python src/watchlist_sync.py

Erforderliche ENVs (gleich wie marketdata_sync.py):
    GDRIVE_SA_KEY     — Service-Account-JSON
    STATE_DOC_ID      — File-ID des STATE-Docs
    JOURNAL_PARENT_ID — Folder-ID, in dem das Journal liegt
                        (default: gleich BRIEFING_FOLDER_ID-Parent, also
                        Workspace-Shared-Drive Trading-Pipeline-Root)

Schema-Stand: 2026-05-23 v2 — 3-Trigger-Schema mit 🚦-Ampel.
    Das Watchlist-Sheet hat 12 Spalten:
        Aktie | Symbol | Richtung | 🚦 A | Trigger A | 🚦 B | Trigger B
              | 🚦 C | Trigger C | Bemerkungen | Datum hinzugefügt | Verfallsdatum
    Jeder der drei Trigger-Slots ist optional und trägt ein eigenes
    Ampel-Emoji (🟢 scharf / 🟡 beobachten / ⏳ wartet / 🔴 tot). Die drei
    Slots werden als ein Trigger-Block `🟢 A) … · 🔴 B) …` in die (unveränderte)
    6-spaltige STATE-Doc-Tabelle gerendert; state_parser splittet sie wieder
    auf, filter_engine überspringt 🔴/⏳-Trigger. Die alte separate
    These-/Status-Spalte ist entfallen — die STATE-Status-Spalte ist fix
    `⚠️ aktiv`, das Datum-Constraint kommt weiterhin aus `nach JJJJ-MM-TT`
    im Trigger-Text.
"""

from __future__ import annotations

import io
import logging
import os
import re
import sys
from datetime import date, datetime
from typing import Optional

from googleapiclient.errors import HttpError
from googleapiclient.http import MediaInMemoryUpload, MediaIoBaseDownload
from openpyxl import load_workbook

# Module aus diesem Repo
from drive_writer import build_drive_service

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("watchlist_sync")


# ===========================================================================
# Excel → Watchlist-Einträge
# ===========================================================================

# Erlaubte 🚦-Ampel-Emojis (Reihenfolge entspricht filter_engine-Semantik):
# 🟢 scharf · 🟡 beobachten · ⏳ wartet (Datum/Bedingung) · 🔴 tot.
VALID_GATES: tuple[str, ...] = ("🟢", "🟡", "⏳", "🔴")

# Default-Gate, wenn eine Gate-Zelle leer oder unlesbar ist: 🟢 (scharf).
# Bewusst der "wird-ausgewertet"-Wert — ein Trigger ohne erkennbares Gate soll
# nicht stillschweigend übersprungen werden (das wäre 🔴/⏳).
_DEFAULT_GATE = "🟢"


def _normalize_gate(raw: str) -> str:
    """Extrahiert das 🚦-Ampel-Emoji aus einer Gate-Zelle.

    Akzeptiert 🟢/🟡/⏳/🔴 (auch wenn die Zelle zusätzlichen Text enthält).
    Leere oder unbekannte Zelle → `_DEFAULT_GATE` (🟢).
    """
    raw = (raw or "").strip()
    for gate in VALID_GATES:
        if gate in raw:
            return gate
    return _DEFAULT_GATE


def _format_trigger_for_state(label: str, gate: str, text: str) -> str:
    """Rendert einen Trigger-Slot als `{gate} {label}) {content}` für den
    STATE-Doc-Trigger-Block.

    Ein im Zelltext bereits vorhandener `{label})`/`{label}:`-Marker (Artefakt
    der Schema-Migration, z.B. "A) Daily-Close …") wird einmalig entfernt,
    damit der Marker nicht doppelt erscheint und das Gate-Emoji direkt vor dem
    Marker sitzt — so bleibt der state_parser-Splitter eindeutig.
    Interne Zeilenumbrüche werden zu Leerzeichen (Markdown-Tabellenzelle ist
    einzeilig). Leerer Text → leerer String (Slot wird vom Aufrufer übersprungen).
    """
    text = (text or "").strip()
    if not text:
        return ""
    marker = re.compile(rf"(?<![A-Za-z0-9]){re.escape(label)}[\)\:]\s*")
    text = marker.sub("", text, count=1).strip()
    text = re.sub(r"\s*\n\s*", " ", text)
    return f"{gate} {label}) {text}"


def _shorten_richtung(richtung: str) -> str:
    """Excel-Richtung kann Klammer-Modifier enthalten (z.B. 'LONG (Pullback)').

    Pipeline parst nur LONG/SHORT — wir nehmen das erste Wort.
    """
    if not richtung:
        return ""
    s = str(richtung).strip()
    first = s.split()[0].upper().rstrip(",.")
    if first in ("LONG", "SHORT"):
        return first
    if "LONG" in s.upper():
        return "LONG"
    if "SHORT" in s.upper():
        return "SHORT"
    return first  # KONDITIONAL etc. unverändert lassen


_VERFALL_DATE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")


def _parse_verfall_date(raw: str) -> str:
    """Extrahiert das führende TT.MM.JJJJ aus dem Verfallsdatum-Feld (Spalte H)
    und normiert es auf ISO YYYY-MM-DD für den STATE-Doc.

    Das Journal-Feld enthält Datum plus Freitext-Klammer, z.B.
    '11.06.2026 (Momentum-Setup +14 HT)'. Leerer String bei fehlendem oder
    unparsbarem Datum.
    """
    if not raw:
        return ""
    m = _VERFALL_DATE_RE.search(raw)
    if not m:
        return ""
    d, mo, y = m.groups()
    try:
        return date(int(y), int(mo), int(d)).isoformat()
    except ValueError:
        return ""


def read_journal_watchlist(xlsx_bytes: bytes) -> list[dict]:
    """Liest das Watchlist-Sheet aus dem Journal und gibt Dicts zurück.

    Erwartet das 3-Trigger-Schema (ab 2026-05-23) mit den Spalten:
        Aktie | Symbol | Richtung | 🚦 A | Trigger A | 🚦 B | Trigger B
              | 🚦 C | Trigger C | Bemerkungen | Datum hinzugefügt | Verfallsdatum

    Jeder Eintrag bekommt eine `triggers`-Liste mit bis zu drei Slots, je ein
    Dict {label, gate, text}. Leere Trigger-Slots werden ausgelassen. Symbol
    und mindestens ein Trigger sind Pflicht — Zeilen ohne werden mit Warnung
    übersprungen.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    if "Watchlist" not in wb.sheetnames:
        raise RuntimeError(
            f"Watchlist-Sheet nicht gefunden im Journal. "
            f"Vorhandene Sheets: {wb.sheetnames}"
        )
    ws = wb["Watchlist"]

    # Header-Zeile suchen — erste Zeile mit dem Wort "Aktie" oder "Symbol"
    header_row_idx = None
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        cells = [str(c).strip().lower() if c else "" for c in row]
        if any("aktie" in c or "symbol" in c for c in cells):
            header_row_idx = i
            headers = cells
            break
    if header_row_idx is None:
        raise RuntimeError("Header-Zeile in Watchlist-Sheet nicht gefunden")

    # Spalten-Indizes finden — erster Header, der einen Kandidaten als
    # Substring enthält.
    def find_col(*candidates: str) -> Optional[int]:
        for cand in candidates:
            for j, h in enumerate(headers):
                if cand in h:
                    return j
        return None

    col_aktie = find_col("aktie")
    col_symbol = find_col("symbol")
    col_richtung = find_col("richtung")
    col_trigger = {
        "A": find_col("trigger a"),
        "B": find_col("trigger b"),
        "C": find_col("trigger c"),
    }
    # Gate-Spalten: Header sind "🚦 A" etc. — "🚦" als Anker, Fallback-Aliase
    # für den Fall, dass das Ampel-Symbol mal anders heißt.
    col_gate = {
        "A": find_col("🚦 a", "🚦a", "ampel a", "gate a"),
        "B": find_col("🚦 b", "🚦b", "ampel b", "gate b"),
        "C": find_col("🚦 c", "🚦c", "ampel c", "gate c"),
    }
    col_verfall = find_col("verfallsdatum", "verfall")

    if col_symbol is None:
        raise RuntimeError(
            "Symbol-Spalte nicht im Watchlist-Sheet gefunden. "
            "Bitte Spalte 'Symbol' anlegen mit Yahoo-Tickern (z.B. CBK.DE, AAPL)."
        )
    if col_aktie is None or col_richtung is None:
        raise RuntimeError(
            "Pflichtspalten 'Aktie' oder 'Richtung' nicht gefunden."
        )
    if col_trigger["A"] is None:
        raise RuntimeError(
            "Spalte 'Trigger A' nicht gefunden — das Watchlist-Sheet scheint "
            "noch im alten Schema zu sein. watchlist_sync erwartet seit "
            "2026-05-23 das 3-Trigger-Schema (Aktie|Symbol|Richtung|🚦 A|"
            "Trigger A|🚦 B|Trigger B|🚦 C|Trigger C|Bemerkungen|"
            "Datum hinzugefügt|Verfallsdatum)."
        )

    entries: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue

        def cell(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        aktie = cell(col_aktie)
        symbol = cell(col_symbol)
        if not aktie or not symbol:
            if aktie or symbol:
                logger.warning(
                    "Zeile übersprungen — Aktie oder Symbol fehlt: "
                    "aktie=%r symbol=%r", aktie, symbol,
                )
            continue

        triggers: list[dict] = []
        for slot in ("A", "B", "C"):
            text = cell(col_trigger[slot])
            if not text:
                continue
            triggers.append({
                "label": slot,
                "gate": _normalize_gate(cell(col_gate[slot])),
                "text": text,
            })
        if not triggers:
            logger.warning(
                "Zeile übersprungen — kein Trigger gesetzt: symbol=%r", symbol,
            )
            continue

        entries.append({
            "aktie": aktie,
            "symbol": symbol,
            "richtung": _shorten_richtung(cell(col_richtung)),
            "triggers": triggers,
            "verfall": _parse_verfall_date(cell(col_verfall)),
        })
    return entries


# ===========================================================================
# Markdown-Block rendern
# ===========================================================================

WATCHLIST_HEADER = "## Watchlist-Trigger (aktive Einträge)"
PFLEGE_BLOCK = """\
> **Pflege:** Tot-Einträge (alle Trigger 🔴 / These geplatzt / gelaufen)
> aus dem Watchlist-Sheet ins "Watchlist-Archiv"-Sheet verschieben.
> Pipeline ignoriert archivierte Einträge automatisch.
>
> **🚦-Ampel pro Trigger (A/B/C) — steht als Präfix im Trigger-Feld:**
> - 🟢 scharf — Trigger wird ausgewertet, Pipeline meldet bei Reichweite
> - 🟡 beobachten — Trigger wird ausgewertet, passiv beobachtet
> - ⏳ wartet — Datum/Bedingung noch nicht erreicht, von der Pipeline übersprungen
> - 🔴 tot — Trigger durchgelaufen/invalidiert, von der Pipeline übersprungen
>
> Die Status-Spalte ist eintragsweit fix `⚠️ aktiv` — die handelbare
> Differenzierung liegt in der 🚦-Ampel je Trigger. Das Datum-Constraint
> wird weiterhin aus `nach JJJJ-MM-TT` im Trigger-Text gelesen.
>
> **AUTO-SYNC:** Diese Tabelle wird bei jedem Tier-A-Lauf (alle 30 Min) aus
> dem Watchlist-Sheet des Trading-Journals neu generiert. Manuelle Edits
> hier werden überschrieben — bitte Watchlist im Journal pflegen."""


def render_watchlist_block(entries: list[dict], generated_at: datetime) -> str:
    """Generiert den kompletten Watchlist-Block (Header + Pflege + Tabelle).

    Die Tabelle hat unverändert 6 Spalten; das Trigger-Feld bündelt die bis
    zu drei Trigger-Slots als `🟢 A) … · 🔴 B) …`. Die Status-Spalte ist
    eintragsweit fix `⚠️ aktiv`.
    """
    lines = [WATCHLIST_HEADER, "", PFLEGE_BLOCK, ""]
    lines.append(
        f"> _Auto-Sync zuletzt: {generated_at.strftime('%Y-%m-%d %H:%M UTC')} "
        f"({len(entries)} Einträge)_"
    )
    lines.append("")
    lines.append("| Kandidat | Symbol | Richtung | Trigger (🚦 A/B/C) | Status | Verfall |")
    lines.append("|----------|--------|----------|--------------------|--------|---------|")
    for e in entries:
        # Trigger-Block: alle nicht-leeren Slots als '{gate} {label}) {text}',
        # mit ' · ' verbunden — state_parser splittet daran wieder auf.
        trigger_cell = " · ".join(
            s for s in (
                _format_trigger_for_state(t["label"], t["gate"], t["text"])
                for t in e["triggers"]
            ) if s
        )
        # Pipe in Zellinhalten escapen, sonst kaputte Tabelle
        kandidat = e["aktie"].replace("|", "\\|")
        trigger_cell = trigger_cell.replace("|", "\\|").replace("\n", " ")
        verfall = e.get("verfall", "").replace("|", "\\|")
        lines.append(
            f"| {kandidat} | {e['symbol']} | {e['richtung']} | {trigger_cell} | "
            f"⚠️ aktiv | {verfall} |"
        )
    lines.append("")
    return "\n".join(lines)


# ===========================================================================
# STATE-Doc Read / Write
# ===========================================================================

# Block-Ersetzung: alles zwischen WATCHLIST_HEADER und der nächsten Sektion
_WATCHLIST_BLOCK_RE = re.compile(
    r"(?ms)^##\s*Watchlist-Trigger.*?(?=^##\s|\Z)"
)


def replace_watchlist_block(state_text: str, new_block: str) -> str:
    """Ersetzt den bestehenden Watchlist-Block im STATE-Text.

    Findet den Block über das Header-Pattern und ersetzt bis zum nächsten
    `## ` oder Dateiende. Wenn kein Block existiert, wird der neue Block
    direkt nach `# STATE START` (oder am Anfang) eingefügt.
    """
    new_block_normalized = new_block.rstrip() + "\n\n"
    if _WATCHLIST_BLOCK_RE.search(state_text):
        return _WATCHLIST_BLOCK_RE.sub(new_block_normalized, state_text, count=1)
    # Kein Block gefunden — nach STATE-START einfügen
    start_match = re.search(r"#\s*STATE\s*START\s*\n", state_text, re.IGNORECASE)
    if start_match:
        idx = start_match.end()
        return state_text[:idx] + "\n" + new_block_normalized + state_text[idx:]
    # Notfall: ganz vorne anhängen
    return new_block_normalized + state_text


def read_state_doc(drive_service, state_doc_id: str) -> tuple[str, str]:
    """Returnt (text, mime_type)."""
    from drive_writer import _with_retry  # lokaler Import, Zyklusvermeidung
    metadata = _with_retry(
        "read_state_doc.metadata",
        lambda: drive_service.files().get(
            fileId=state_doc_id, fields="mimeType,name", supportsAllDrives=True,
        ).execute(),
    )
    mime_type = metadata.get("mimeType", "")

    if mime_type == "application/vnd.google-apps.document":
        request = drive_service.files().export_media(
            fileId=state_doc_id, mimeType="text/plain",
        )
    else:
        request = drive_service.files().get_media(
            fileId=state_doc_id, supportsAllDrives=True,
        )
    content = _with_retry("read_state_doc.content", lambda: request.execute())
    if isinstance(content, bytes):
        for enc in ("utf-8", "cp1252", "latin-1"):
            try:
                return content.decode(enc), mime_type
            except UnicodeDecodeError:
                continue
    return str(content), mime_type


def write_state_doc(drive_service, state_doc_id: str, new_text: str, mime_type: str) -> None:
    """Schreibt den neuen Inhalt zurück.

    mime_type-aware (gefixt 2026-05-19, Note #68):
    - Bei Google Docs Native (`vnd.google-apps.document`): Upload als
      `text/plain`. Drive akzeptiert `text/markdown` per files().update()
      mit `uploadType=media` unzuverlässig — beobachtet HTTP 400 Bad Request
      seit ~Mitte Mai 2026, Folge: stiller Sync-Drift (CHKP-Schaden 18.05.).
      Plain-Text-Upload behält die Markdown-Syntax (## Headers, | Tabellen)
      als Klartext im Doc — state_parser liest via text/plain-Export sauber
      zurück. Im Doc-Editor sieht's roh aus, das ist akzeptabel weil der
      STATE-Doc primär von der Pipeline gelesen wird, nicht von Menschen.
    - Bei Markdown-File (`text/markdown` o.ä.): Original-mime_type beibehalten.
    """
    from drive_writer import _with_retry  # lokaler Import

    if mime_type == "application/vnd.google-apps.document":
        upload_mime = "text/plain"
    elif mime_type:
        upload_mime = mime_type
    else:
        upload_mime = "text/markdown"

    logger.info(
        "write_state_doc: target_mime=%s, upload_mime=%s, bytes=%d",
        mime_type or "unknown", upload_mime, len(new_text.encode("utf-8")),
    )

    media = MediaInMemoryUpload(
        new_text.encode("utf-8"),
        mimetype=upload_mime,
        resumable=False,
    )
    _with_retry(
        "write_state_doc",
        lambda: drive_service.files().update(
            fileId=state_doc_id,
            media_body=media,
            supportsAllDrives=True,
        ).execute(),
    )


# ===========================================================================
# Journal-Suche
# ===========================================================================

def find_latest_journal(drive_service, search_folder_id: Optional[str] = None) -> dict:
    """Sucht jüngste Trading_Journal_*.xlsx-Datei.

    Suche zuerst im angegebenen Folder (falls gesetzt), dann global im
    Account. Returnt Dict mit id, name, modifiedTime.
    """
    queries: list[str] = []
    if search_folder_id:
        queries.append(
            f"'{search_folder_id}' in parents "
            f"and name contains 'Trading_Journal' "
            f"and trashed = false"
        )
    queries.append(
        "name contains 'Trading_Journal' "
        "and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
        "and trashed = false"
    )

    for q in queries:
        try:
            from drive_writer import _with_retry  # lokaler Import
            result = _with_retry(
                "find_latest_journal.list",
                lambda q=q: drive_service.files().list(
                    q=q,
                    orderBy="modifiedTime desc",
                    fields="files(id,name,modifiedTime,parents)",
                    pageSize=10,
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True,
                    corpora="allDrives",
                ).execute(),
            )
            files = result.get("files", [])
            if files:
                logger.info(
                    "Journal gefunden: %s (modifiziert %s)",
                    files[0]["name"], files[0]["modifiedTime"],
                )
                return files[0]
        except HttpError as e:
            logger.warning("Drive-Search fehlgeschlagen: %s", e)
            continue

    raise RuntimeError(
        "Kein Trading_Journal_*.xlsx im Drive gefunden. "
        "Bitte Journal in den Workspace-Shared-Drive 'Trading-Pipeline' "
        "hochladen oder via Drive Desktop syncen."
    )


def download_journal(drive_service, file_id: str) -> bytes:
    """Lädt Journal als Bytes runter."""
    from drive_writer import _with_retry  # lokaler Import
    request = drive_service.files().get_media(
        fileId=file_id, supportsAllDrives=True,
    )
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        # next_chunk() macht intern auch HTTPs — bei TLS-Reset abfangen.
        _, done = _with_retry(
            "download_journal.next_chunk",
            lambda: downloader.next_chunk(),
        )
    return buf.getvalue()


# ===========================================================================
# Main
# ===========================================================================

def main() -> int:
    state_doc_id = os.environ.get("STATE_DOC_ID")
    journal_parent_id = os.environ.get("JOURNAL_PARENT_ID")  # optional

    if not state_doc_id:
        logger.error("STATE_DOC_ID env variable nicht gesetzt")
        return 1

    drive_service = build_drive_service()

    # 1. Journal finden & laden
    journal_meta = find_latest_journal(drive_service, journal_parent_id)
    journal_bytes = download_journal(drive_service, journal_meta["id"])

    # 2. Watchlist-Sheet parsen
    entries = read_journal_watchlist(journal_bytes)
    logger.info("Watchlist aus Journal: %d Einträge", len(entries))
    if not entries:
        logger.warning("Watchlist leer — Sync abgebrochen, STATE bleibt unverändert")
        return 0

    # 3. STATE-Doc lesen
    state_text, mime_type = read_state_doc(drive_service, state_doc_id)

    # 3a. Markdown-Escapes strippen (Note #68 Round 2, 2026-05-19):
    # Drive's text/plain-Export escaped Markdown-Sonderzeichen mit Backslash.
    # Ohne Strip akkumulieren bei jedem Read/Write-Cycle weitere Backslashes,
    # bis das Doc exponentiell aufbläht und Drive HTTP 400 zurückgibt.
    # state_parser hatte den Strip bereits beim Read — wir müssen ihn auch
    # hier anwenden, weil watchlist_sync das Doc nicht nur liest sondern
    # zurückschreibt.
    from state_parser import _strip_markdown_escapes
    original_len = len(state_text)
    state_text = _strip_markdown_escapes(state_text)
    if len(state_text) < original_len:
        logger.info(
            "Markdown-Escapes gestrippt: %d → %d Zeichen (%.1f%% Reduktion)",
            original_len, len(state_text),
            (original_len - len(state_text)) / original_len * 100,
        )

    # 4. Watchlist-Block neu rendern und ersetzen
    new_block = render_watchlist_block(entries, datetime.utcnow())
    new_text = replace_watchlist_block(state_text, new_block)

    if new_text == state_text:
        logger.info("Watchlist-Block unverändert — kein Update nötig")
        return 0

    # 5. STATE-Doc schreiben
    write_state_doc(drive_service, state_doc_id, new_text, mime_type)
    logger.info("STATE-Doc aktualisiert (%d Einträge synchronisiert)", len(entries))

    # 6. Sanity-Check (Note #68): Re-Read und Eintrags-Count vergleichen.
    # Drive-API kann Markdown-Uploads stumm verlieren oder konvertieren —
    # wenn das passiert, ist der Sync gescheitert, auch wenn files().update()
    # 200 OK lieferte. Lieber rote Pipeline als grüne Lüge.
    state_text_after, _ = read_state_doc(drive_service, state_doc_id)
    state_text_after = _strip_markdown_escapes(state_text_after)
    after_count = _count_watchlist_entries(state_text_after)
    if after_count != len(entries):
        logger.warning(
            "STATE-Doc-Sanity-Check: erwartet %d WL-Einträge nach Write, "
            "gefunden %d. Sync NICHT abgebrochen (additive Intention — Tier-A "
            "muss weiterlaufen). Manuell prüfen: STATE_DOC_ID=%s",
            len(entries), after_count, state_doc_id,
        )
    else:
        logger.info(
            "Sanity-Check OK: %d Einträge im STATE-Doc bestätigt", after_count
        )
    return 0


def _count_watchlist_entries(state_text: str) -> int:
    """Zählt Datenzeilen in der Watchlist-Tabelle des STATE-Docs.

    Heuristik: nach dem WATCHLIST_HEADER alle Zeilen, die mit `|` beginnen,
    keine Separator-/Header-Zeilen sind, und nicht in einer Nachbar-Sektion
    liegen. Robust gegen text/plain-Export aus Google Docs (`##`-Headers
    bleiben als Plain-Text erhalten).
    """
    match = _WATCHLIST_BLOCK_RE.search(state_text)
    if not match:
        return 0
    block = match.group(0)
    count = 0
    for line in block.splitlines():
        s = line.strip()
        if not s.startswith("|"):
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if len(cells) >= 2 and cells[0] == "Kandidat" and cells[1] == "Symbol":
            continue  # Header-Zeile (nur echte Kopfzeile, nicht Trigger-Text)
        if re.match(r"^\|\s*[-:]+\s*\|", s):
            continue  # Separator
        count += 1
    return count


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        logger.exception("Watchlist-Sync fehlgeschlagen: %s", e)
        sys.exit(1)
