"""
journal_utils.py — Helper-Modul für das Trading-Journal von Matthias

Bündelt alle wiederkehrenden openpyxl-Operationen: Trade eintragen, schließen,
Teilexit, Watchlist, Übersicht-Sync. Jede Operation hält alle abhängigen
Sheets konsistent (Detail-Sheet + SALDO + Geschlossene Trades + Übersicht).

Verwendung (Beispiel — Trade schließen):

    import journal_utils as ju

    wb = ju.open_journal('Trading_Journal_20260417.xlsx')
    ju.close_derivate_trade(
        wb, nr=50, verkaufsdatum='18.04.2026', erloes=543.90,
        lektion='SL bei 0,74€ ausgelöst, Underlying sprang über KO-Puffer'
    )
    ju.update_timestamp(wb, grund='HDD Short ausgestoppt')
    ju.save_journal(wb, 'Trading_Journal_20260418.xlsx')

Stand: 24.04.2026 — Journal-Layout v3 (Gebühren-Transparenz: SK Spalten N+O,
AV Spalten O+P, Archiv Spalten M+N, Übersicht Gebühren-Summenzeile R15).
"""

from __future__ import annotations

from copy import copy
from typing import Optional, List, Dict, Any, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# KONSTANTEN — Sheet-Namen und Styling
# ============================================================

SHEET_UEBERSICHT = 'Übersicht'
SHEET_WATCHLIST = 'Watchlist'
SHEET_GESCHLOSSEN = 'Geschlossene Trades'
SHEET_SK = 'Sonstige Kapitalerträge'          # Derivate + ETFs + ETPs + ETCs
SHEET_AV = 'Aktienveräußerungen'              # Direktaktien (separater Topf!)
SHEET_WK = 'Werbungskosten'
SHEET_NOTES = 'Notes'                         # Handlungsbedarf + Milestones + ⚠️

# Portfolio-Block Konstanten (Übersicht)
PORTFOLIO_HEADER_ROW = 17
PORTFOLIO_START_ROW = 18            # erste Datenzeile

# Fills
FILL_GELB = PatternFill('solid', fgColor='FFFFFF00')                # OFFEN-Markierung
FILL_NONE = PatternFill(fill_type=None)
FILL_ZEBRA = PatternFill('solid', fgColor='FFF5F5F5')               # Zebra-Streifen Portfolio
FILL_HEADER_BLAU = PatternFill('solid', fgColor='FF1B3A5C')         # Header dunkelblau
FILL_SUMME_GRUEN = PatternFill('solid', fgColor='FFE2EFDA')         # SUMME OFFEN / GESAMT

# Farben
COLOR_GEWINN = 'FF1B7A2B'       # dunkelgrün
COLOR_VERLUST = 'FFCC0000'      # kräftiges rot
COLOR_HEADER_WEISS = 'FFFFFFFF'

# Fonts
FONT_CALIBRI = Font(name='Calibri', size=11, bold=False)
FONT_CALIBRI_BOLD = Font(name='Calibri', size=11, bold=True)
FONT_CAMBRIA = Font(name='Cambria', size=11, bold=False)
FONT_GV_GEWINN = Font(name='Cambria', size=11, bold=False, color=COLOR_GEWINN)
FONT_GV_VERLUST = Font(name='Cambria', size=11, bold=False, color=COLOR_VERLUST)
FONT_SALDO_GRUEN = Font(name='Cambria', size=12, bold=True, color=COLOR_GEWINN)
FONT_SALDO_ROT = Font(name='Cambria', size=12, bold=True, color=COLOR_VERLUST)
FONT_STEUER = Font(name='Cambria', size=11, bold=False, color=COLOR_VERLUST)
FONT_NETTO_GRUEN = Font(name='Cambria', size=11, bold=True, color=COLOR_GEWINN)
FONT_NETTO_ROT = Font(name='Cambria', size=11, bold=True, color=COLOR_VERLUST)
FONT_SALDO_HEADER = Font(name='Arial', size=10, bold=True)

# Number-Formats
FMT_EURO = '#,##0.00" €"'
FMT_PROZENT = '0.0%'
FMT_GENERAL = 'General'

# Saldo-Labels (in Spalte A)
LABEL_SALDO = 'REALISIERTER SALDO'
LABEL_STEUER = 'STEUER (26,375%)'
LABEL_NETTO = 'REINGEWINN (NETTO)'

# Abgeltungssteuer
STEUERSATZ = 0.26375


# ============================================================
# OPEN / SAVE / RELOAD
# ============================================================

def open_journal(path: str) -> Workbook:
    """Lädt das Journal. Gibt ein Workbook-Objekt zurück."""
    return load_workbook(path)


def save_journal(wb: Workbook, output_path: str) -> Workbook:
    """
    Speichert das Journal und lädt es neu (Ghost-Value-Workaround).
    Immer das zurückgegebene Workbook für weitere Änderungen nutzen.
    """
    wb.save(output_path)
    return load_workbook(output_path)


def reload_journal(path: str) -> Workbook:
    """Explizites Neu-Laden — z.B. nach externem Schreibvorgang."""
    return load_workbook(path)


# ============================================================
# STYLE-HELPERS
# ============================================================

def _copy_style(src, dst):
    """Kopiert Font/Border/Alignment/Fill/Format von src-Zelle auf dst-Zelle."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _set_gelb(ws: Worksheet, row: int, col_start: int = 1, col_end: int = 13):
    """Markiert eine Zeile gelb (OFFEN)."""
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = FILL_GELB


def _remove_gelb(ws: Worksheet, row: int, col_start: int = 1, col_end: int = 13):
    """Entfernt gelbe Markierung (Position geschlossen)."""
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = FILL_NONE


def _gv_font(gv_value: float) -> Font:
    """Wählt die richtige Schriftfarbe für G/V: grün bei Gewinn, rot bei Verlust."""
    return FONT_GV_GEWINN if gv_value >= 0 else FONT_GV_VERLUST


# ============================================================
# FINDER-FUNKTIONEN
# ============================================================

def find_saldo_rows(ws: Worksheet) -> Dict[str, int]:
    """
    Findet die drei Saldo-Zeilen im SK-Sheet über die Label in Spalte A.
    Liefert Dict mit Keys 'saldo', 'steuer', 'netto'.
    """
    result = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val == LABEL_SALDO:
            result['saldo'] = row
        elif val == LABEL_STEUER:
            result['steuer'] = row
        elif val == LABEL_NETTO:
            result['netto'] = row
    if len(result) != 3:
        raise ValueError(
            f"Saldo-Block nicht vollständig gefunden: {result}. "
            f"Erwarte drei Zeilen mit Labels in Spalte A."
        )
    return result


def find_next_trade_nr(ws: Worksheet) -> int:
    """
    Nächste freie Trade-Nummer = max(existierende Integer-Nr) + 1.
    Ignoriert '—'-Zeilen (Steuerkorrekturen) und leere Zellen.
    """
    max_nr = 0
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, int):
            max_nr = max(max_nr, val)
    return max_nr + 1


def find_trade_row(ws: Worksheet, nr: int, status_filter: Optional[str] = None) -> Optional[int]:
    """
    Findet die Zeile eines Trades per Nr. Status-Spalte hängt vom Sheet ab:
    SK = Spalte 12 (L), AV = Spalte 13 (M).
    Wenn mehrere Zeilen dieselbe Nr haben (TEIL-EXIT), liefert die erste mit
    passendem Status oder — wenn nicht gefiltert — die erste überhaupt.
    """
    sheet_name = ws.title
    if sheet_name == SHEET_SK:
        status_col = 12
    elif sheet_name == SHEET_AV:
        status_col = 13
    else:
        raise ValueError(f"Unbekanntes Sheet: {sheet_name}")

    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val == nr:
            if status_filter is None:
                return row
            if ws.cell(row=row, column=status_col).value == status_filter:
                return row
    return None


def find_all_trade_rows(ws: Worksheet, nr: int) -> List[int]:
    """Alle Zeilen mit gegebener Trade-Nr (für TEIL-EXIT-Handling)."""
    rows = []
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == nr:
            rows.append(row)
    return rows


def find_summe_offen_row(ws_uebersicht: Worksheet) -> Optional[int]:
    """Findet die Zeile 'SUMME OFFEN:' im Portfolio-Block der Übersicht."""
    for row in range(PORTFOLIO_START_ROW, ws_uebersicht.max_row + 2):
        val = ws_uebersicht.cell(row=row, column=1).value
        if isinstance(val, str) and val.startswith('SUMME OFFEN'):
            return row
    return None


# ============================================================
# SALDO-BLOCK AKTUALISIEREN (SK Sheet)
# ============================================================

def _update_saldo_block(ws: Worksheet):
    """
    Rechnet den Saldo-Block am Ende von "Sonstige Kapitalerträge" neu:
    - Summiert alle G/V-Werte in Spalte J aus Zeilen mit Integer-Trade-Nr
    - Aktualisiert Gewinne-String (H), Verluste-String (I), Saldo (J)
    - Berechnet Steuer (26,375% × Saldo, nur bei Gewinn)
    - Berechnet Netto (Saldo - Steuer)
    - Setzt korrekte Fonts (grün für Saldo/Netto, rot für Steuer)
    """
    rows = find_saldo_rows(ws)
    saldo_row = rows['saldo']
    steuer_row = rows['steuer']
    netto_row = rows['netto']

    gewinne = 0.0
    verluste = 0.0
    for r in range(5, saldo_row):
        nr = ws.cell(row=r, column=1).value
        gv = ws.cell(row=r, column=10).value  # Spalte J
        if isinstance(nr, int) and isinstance(gv, (int, float)):
            if gv > 0:
                gewinne += gv
            elif gv < 0:
                verluste += abs(gv)

    saldo = gewinne - verluste
    steuer = -saldo * STEUERSATZ if saldo > 0 else 0.0
    netto = saldo + steuer

    # Strings H/I der Saldo-Zeile
    ws.cell(row=saldo_row, column=8).value = f"Gewinne: {gewinne:.2f}€"
    ws.cell(row=saldo_row, column=9).value = f"Verluste: {verluste:.2f}€"

    # Zahlenwerte
    c_saldo = ws.cell(row=saldo_row, column=10)
    c_saldo.value = round(saldo, 2)
    c_saldo.number_format = FMT_EURO
    c_saldo.font = FONT_SALDO_GRUEN if saldo >= 0 else FONT_SALDO_ROT

    c_steuer = ws.cell(row=steuer_row, column=10)
    c_steuer.value = round(steuer, 2)
    c_steuer.number_format = FMT_EURO
    c_steuer.font = FONT_STEUER

    c_netto = ws.cell(row=netto_row, column=10)
    c_netto.value = round(netto, 2)
    c_netto.number_format = FMT_EURO
    c_netto.font = FONT_NETTO_GRUEN if netto >= 0 else FONT_NETTO_ROT


# ============================================================
# SK SHEET — DERIVATE-TRADE EINTRAGEN
# ============================================================

def add_derivate_trade(wb: Workbook, trade: Dict[str, Any]) -> int:
    """
    Trägt einen neuen Derivate-Trade in "Sonstige Kapitalerträge" ein.

    Erwartete Keys in `trade` (Pflicht fett):
      **kaufdatum**      : 'DD.MM.YYYY' (String!)
      **instrument**     : Name z.B. 'HSBC OE Turbo Long CTS Eventim'
      **isin**           : ISIN oder WKN (ggf. mit ⚠️-Präfix bei Unsicherheit)
      **typ**            : 'KO Long' | 'KO Short' | 'ETF' | 'ETP' | 'ETC'
      **richtung**       : 'Long' | 'Short'
      **kaufsumme**      : float — Einsatz in €
      **notizen**        : String mit Stk/Kurs/SL/TP/These

      nr                 : int — wenn weggelassen, automatisch nächste Nr
      status             : default 'OFFEN'

    Return: Trade-Nr der eingetragenen Position.
    """
    ws = wb[SHEET_SK]
    rows = find_saldo_rows(ws)
    saldo_row = rows['saldo']

    nr = trade.get('nr') or find_next_trade_nr(ws)

    # Zeile vor der SALDO-Zeile einfügen
    new_row = saldo_row
    ws.insert_rows(new_row)

    # Befüllen
    _fill_derivate_row(ws, new_row, nr=nr, trade=trade, status=trade.get('status', 'OFFEN'))

    # Gelb markieren wenn OFFEN
    if trade.get('status', 'OFFEN') == 'OFFEN':
        _set_gelb(ws, new_row, 1, 13)

    return nr


def _fill_derivate_row(ws: Worksheet, row: int, nr: int, trade: Dict[str, Any], status: str):
    """Schreibt Derivate-Trade-Daten in eine Zeile. Setzt konsistente Formate."""
    ws.cell(row=row, column=1).value = nr                       # A: Nr
    ws.cell(row=row, column=2).value = trade['kaufdatum']       # B: Kaufdatum
    ws.cell(row=row, column=3).value = trade.get('verkaufsdatum')  # C
    ws.cell(row=row, column=4).value = trade['instrument']      # D
    ws.cell(row=row, column=5).value = trade['isin']            # E
    ws.cell(row=row, column=6).value = trade['typ']             # F
    ws.cell(row=row, column=7).value = trade['richtung']        # G

    # Kaufsumme H
    c_h = ws.cell(row=row, column=8)
    c_h.value = float(trade['kaufsumme'])
    c_h.number_format = FMT_EURO

    # Erlös I (bei OFFEN leer)
    c_i = ws.cell(row=row, column=9)
    erloes = trade.get('erloes')
    if erloes is not None:
        c_i.value = float(erloes)
    c_i.number_format = FMT_EURO

    # G/V J
    c_j = ws.cell(row=row, column=10)
    if erloes is not None:
        gv = float(erloes) - float(trade['kaufsumme'])
        c_j.value = round(gv, 2)
        c_j.font = _gv_font(gv)
    else:
        c_j.value = 0
    c_j.number_format = FMT_EURO

    # G/V% K
    c_k = ws.cell(row=row, column=11)
    if erloes is not None and float(trade['kaufsumme']) != 0:
        pct = (float(erloes) - float(trade['kaufsumme'])) / float(trade['kaufsumme'])
        c_k.value = pct
        c_k.font = _gv_font(pct)
    c_k.number_format = FMT_PROZENT

    # Status L
    ws.cell(row=row, column=12).value = status

    # Notizen M
    ws.cell(row=row, column=13).value = trade.get('notizen', '')

    # Gebühr Kauf N (Pflicht ab v3 — für Transparenz-Zeile in Übersicht)
    gk = trade.get('gebuehr_kauf')
    if gk is not None:
        c_n = ws.cell(row=row, column=14)
        c_n.value = float(gk)
        c_n.number_format = FMT_EURO

    # Gebühr Verkauf O (nur bei Close befüllt)
    gv_fee = trade.get('gebuehr_verkauf')
    if gv_fee is not None:
        c_o = ws.cell(row=row, column=15)
        c_o.value = float(gv_fee)
        c_o.number_format = FMT_EURO

    # Konsistente Basis-Fonts für Text-Spalten
    for col in [1, 2, 3, 4, 5, 6, 7, 12, 13]:
        cell = ws.cell(row=row, column=col)
        if cell.font.color is None or not isinstance(cell.font.color.rgb, str):
            cell.font = FONT_CALIBRI


# ============================================================
# SK SHEET — DERIVATE-TRADE SCHLIESSEN
# ============================================================

def close_derivate_trade(
    wb: Workbook,
    nr: int,
    verkaufsdatum: str,
    erloes: float,
    lektion: Optional[str] = None,
    archiv: bool = True,
    gebuehr_verkauf: Optional[float] = None,
) -> int:
    """
    Schließt einen offenen Derivate-Trade komplett.

    1. Findet die OFFEN-Zeile per Nr
    2. Trägt Verkaufsdatum (C), Erlös (I), G/V (J), G/V% (K) ein
    3. Status → 'GESCHLOSSEN'
    4. Entfernt gelbe Markierung
    5. Setzt Font-Farbe auf J/K (grün/rot)
    6. Ergänzt Lektion in Notizen (M)
    7. Aktualisiert SALDO-Block
    8. Trägt Verkauf-Gebühr in O ein (wenn gegeben) — Kauf-Gebühr in N
       bleibt aus der Add-Phase stehen.
    9. Kopiert den geschlossenen Trade ins "Geschlossene Trades"-Archiv (wenn archiv=True)

    Wichtig: erloes sollte die Gebühr bereits abziehen (Erlös netto),
    genau wie kaufsumme beim Entry die Kauf-Gebühr enthält. gebuehr_verkauf
    ist nur die Info-Spalte für die Transparenz-KPI.

    Returns: Die Zeilen-Nr der geschlossenen Zeile.
    """
    ws = wb[SHEET_SK]
    row = find_trade_row(ws, nr, status_filter='OFFEN')
    if row is None:
        raise ValueError(f"Kein OFFENER Trade mit Nr {nr} im Sheet '{SHEET_SK}' gefunden.")

    kaufsumme = float(ws.cell(row=row, column=8).value)
    gv = round(float(erloes) - kaufsumme, 2)
    pct = gv / kaufsumme if kaufsumme != 0 else 0.0

    ws.cell(row=row, column=3).value = verkaufsdatum

    c_i = ws.cell(row=row, column=9)
    c_i.value = round(float(erloes), 2)
    c_i.number_format = FMT_EURO

    c_j = ws.cell(row=row, column=10)
    c_j.value = gv
    c_j.number_format = FMT_EURO
    c_j.font = _gv_font(gv)

    c_k = ws.cell(row=row, column=11)
    c_k.value = pct
    c_k.number_format = FMT_PROZENT
    c_k.font = _gv_font(pct)

    ws.cell(row=row, column=12).value = 'GESCHLOSSEN'

    # Gebühr Verkauf O
    if gebuehr_verkauf is not None:
        c_o = ws.cell(row=row, column=15)
        c_o.value = float(gebuehr_verkauf)
        c_o.number_format = FMT_EURO

    # Notizen erweitern (Lektion anhängen, falls gegeben)
    if lektion:
        existing = ws.cell(row=row, column=13).value or ''
        emoji = '✅' if gv >= 0 else '❌'
        gv_str = f"{gv:+.2f}€ ({pct:+.1%}) {emoji}"
        ws.cell(row=row, column=13).value = f"{existing} | EXIT: {gv_str} | {lektion}".strip(' |')

    _remove_gelb(ws, row, 1, 13)
    _update_saldo_block(ws)

    if archiv:
        _append_to_geschlossene(wb, _collect_sk_row(ws, row))

    return row


def _collect_sk_row(ws: Worksheet, row: int) -> Dict[str, Any]:
    """Sammelt die Daten einer SK-Zeile als Dict (für Archiv-Export)."""
    return {
        'nr': ws.cell(row=row, column=1).value,
        'kaufdatum': ws.cell(row=row, column=2).value,
        'verkaufsdatum': ws.cell(row=row, column=3).value,
        'instrument': ws.cell(row=row, column=4).value,
        'isin': ws.cell(row=row, column=5).value,
        'typ': ws.cell(row=row, column=6).value,
        'richtung': ws.cell(row=row, column=7).value,
        'kaufsumme': ws.cell(row=row, column=8).value,
        'erloes': ws.cell(row=row, column=9).value,
        'gv': ws.cell(row=row, column=10).value,
        'gv_pct': ws.cell(row=row, column=11).value,
        'notizen': ws.cell(row=row, column=13).value,
        'gebuehr_kauf': ws.cell(row=row, column=14).value,
        'gebuehr_verkauf': ws.cell(row=row, column=15).value,
    }


# ============================================================
# SK SHEET — TEIL-EXIT
# ============================================================

def partial_exit_derivate(
    wb: Workbook,
    nr: int,
    verkaufsdatum: str,
    verkaufte_einstand: float,
    verkaufter_erloes: float,
    rest_einstand: float,
    notiz_verkauft: str = '',
    notiz_rest: str = '',
    archiv: bool = True,
) -> Tuple[int, int]:
    """
    Spaltet eine offene Position in TEIL-EXIT (verkaufte Tranche) + neue OFFEN-Zeile (Rest).

    Pattern im Journal (Beispiel Jenoptik #47):
      Z57: Nr=47, Status=TEIL-EXIT, H=anteil. Einstand verkaufter Teil, I=Erlös
      Z58: Nr=47, Status=OFFEN,     H=anteil. Einstand Rest

    Parameter:
      nr                 : Trade-Nr der bestehenden OFFEN-Position
      verkaufsdatum      : 'DD.MM.YYYY'
      verkaufte_einstand : Anteiliger Einstand des verkauften Teils (FIFO)
      verkaufter_erloes  : Erlös des Teilverkaufs
      rest_einstand      : Anteiliger Einstand der Restposition
      notiz_verkauft     : Notiz für die TEIL-EXIT-Zeile
      notiz_rest         : Notiz für die neue OFFEN-Rest-Zeile

    Returns: (teil_exit_row, rest_offen_row)
    """
    ws = wb[SHEET_SK]
    row = find_trade_row(ws, nr, status_filter='OFFEN')
    if row is None:
        raise ValueError(f"Keine OFFENE Position #{nr} gefunden.")

    # Originaldaten sichern, bevor wir die Zeile modifizieren
    orig = _collect_sk_row(ws, row)

    # 1) Bestehende Zeile → TEIL-EXIT mit verkauftem Anteil
    gv = round(verkaufter_erloes - verkaufte_einstand, 2)
    pct = gv / verkaufte_einstand if verkaufte_einstand != 0 else 0.0

    ws.cell(row=row, column=3).value = verkaufsdatum
    c_h = ws.cell(row=row, column=8)
    c_h.value = round(verkaufte_einstand, 2)
    c_h.number_format = FMT_EURO

    c_i = ws.cell(row=row, column=9)
    c_i.value = round(verkaufter_erloes, 2)
    c_i.number_format = FMT_EURO

    c_j = ws.cell(row=row, column=10)
    c_j.value = gv
    c_j.number_format = FMT_EURO
    c_j.font = _gv_font(gv)

    c_k = ws.cell(row=row, column=11)
    c_k.value = pct
    c_k.number_format = FMT_PROZENT
    c_k.font = _gv_font(pct)

    ws.cell(row=row, column=12).value = 'TEIL-EXIT'
    if notiz_verkauft:
        ws.cell(row=row, column=13).value = notiz_verkauft

    # Gelb von dieser Zeile entfernen (sie ist nicht mehr OFFEN)
    _remove_gelb(ws, row, 1, 13)

    # 2) Neue Zeile direkt dahinter für den Rest (OFFEN)
    rest_row = row + 1
    ws.insert_rows(rest_row)

    rest_trade = {
        'kaufdatum': orig['kaufdatum'],
        'instrument': orig['instrument'],
        'isin': orig['isin'],
        'typ': orig['typ'],
        'richtung': orig['richtung'],
        'kaufsumme': rest_einstand,
        'notizen': notiz_rest or f"RESTPOSITION nach TEIL-EXIT {verkaufsdatum}",
    }
    _fill_derivate_row(ws, rest_row, nr=nr, trade=rest_trade, status='OFFEN')
    _set_gelb(ws, rest_row, 1, 13)

    _update_saldo_block(ws)

    if archiv:
        # Nur die verkaufte Tranche ins Archiv
        _append_to_geschlossene(wb, _collect_sk_row(ws, row))

    return row, rest_row



# ============================================================
# AKTIENVERÄUSSERUNGEN — SEPARATER STEUERTOPF
# ============================================================

def _find_av_next_row(ws: Worksheet) -> int:
    """Nächste freie Zeile im Aktienveräußerungen-Sheet (nach dem Datenblock)."""
    # Daten beginnen in Zeile 5. Suche erste Zeile, die komplett leer in A ist.
    for row in range(5, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    return ws.max_row + 1


def _find_av_next_nr(ws: Worksheet) -> int:
    max_nr = 0
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, int):
            max_nr = max(max_nr, val)
    return max_nr + 1


def add_aktie(wb: Workbook, trade: Dict[str, Any]) -> int:
    """
    Trägt einen neuen Direktaktien-Kauf in "Aktienveräußerungen" ein.
    Separater Verrechnungstopf (§20 Abs. 2 Nr. 1 EStG)!

    Pflicht-Keys:
      **kaufdatum**, **aktie**, **isin**, **boerse**, **stueck**,
      **kaufpreis_stk**, **notizen**

      nr (optional — sonst automatisch)
      gebuehr : default 1.0 (TR-Standard)
    """
    ws = wb[SHEET_AV]
    row = _find_av_next_row(ws)
    nr = trade.get('nr') or _find_av_next_nr(ws)
    gebuehr = trade.get('gebuehr', 1.0)

    einsatz = float(trade['stueck']) * float(trade['kaufpreis_stk']) + float(gebuehr)

    ws.cell(row=row, column=1).value = nr
    ws.cell(row=row, column=2).value = trade['kaufdatum']
    ws.cell(row=row, column=4).value = trade['aktie']
    ws.cell(row=row, column=5).value = trade['isin']
    ws.cell(row=row, column=6).value = trade['boerse']
    ws.cell(row=row, column=7).value = trade['stueck']

    c_h = ws.cell(row=row, column=8)
    c_h.value = float(trade['kaufpreis_stk'])
    c_h.number_format = FMT_EURO

    c_j = ws.cell(row=row, column=10)
    c_j.value = round(einsatz, 2)
    c_j.number_format = FMT_EURO

    ws.cell(row=row, column=13).value = 'OFFEN'
    ws.cell(row=row, column=14).value = trade['notizen']

    # Gebühr Kauf O (ab v3 — Transparenz-Spalte; ist zusätzlich schon in einsatz drin)
    c_o = ws.cell(row=row, column=15)
    c_o.value = float(gebuehr)
    c_o.number_format = FMT_EURO

    # Basis-Font für Text-Spalten
    for col in [1, 2, 3, 4, 5, 6, 7, 13, 14]:
        ws.cell(row=row, column=col).font = FONT_CALIBRI

    _set_gelb(ws, row, 1, 16)
    return nr


def close_aktie(
    wb: Workbook,
    nr: int,
    verkaufsdatum: str,
    verkaufspreis_stk: float,
    erloes: Optional[float] = None,
    lektion: Optional[str] = None,
    archiv: bool = True,
    gebuehr_verkauf: Optional[float] = None,
) -> int:
    """
    Schließt eine Aktienposition.
    erloes wird aus Stück × Verkaufspreis abgeleitet, falls nicht gegeben.

    gebuehr_verkauf wird in Spalte P eingetragen (Transparenz) und sollte bereits
    im übergebenen erloes abgezogen sein — sonst übergib erloes mit Gebühr drin.
    """
    ws = wb[SHEET_AV]
    # In AV ist Status in Spalte M (13) — find_trade_row macht das korrekt
    row = find_trade_row(ws, nr, status_filter='OFFEN')
    if row is None:
        raise ValueError(f"Keine OFFENE Aktie #{nr} in '{SHEET_AV}'.")

    stueck = float(ws.cell(row=row, column=7).value)
    einsatz = float(ws.cell(row=row, column=10).value)

    if erloes is None:
        erloes = stueck * float(verkaufspreis_stk)

    gv = round(float(erloes) - einsatz, 2)

    ws.cell(row=row, column=3).value = verkaufsdatum

    c_i = ws.cell(row=row, column=9)
    c_i.value = float(verkaufspreis_stk)
    c_i.number_format = FMT_EURO

    c_k = ws.cell(row=row, column=11)
    c_k.value = round(float(erloes), 2)
    c_k.number_format = FMT_EURO

    c_l = ws.cell(row=row, column=12)
    c_l.value = gv
    c_l.number_format = FMT_EURO
    c_l.font = _gv_font(gv)

    ws.cell(row=row, column=13).value = 'GESCHLOSSEN'

    # Gebühr Verkauf P (Transparenz)
    if gebuehr_verkauf is not None:
        c_p = ws.cell(row=row, column=16)
        c_p.value = float(gebuehr_verkauf)
        c_p.number_format = FMT_EURO

    if lektion:
        existing = ws.cell(row=row, column=14).value or ''
        emoji = '✅' if gv >= 0 else '❌'
        ws.cell(row=row, column=14).value = (
            f"{existing} | EXIT {verkaufsdatum}: {gv:+.2f}€ {emoji} | {lektion}"
        ).strip(' |')

    _remove_gelb(ws, row, 1, 16)

    if archiv:
        _append_to_geschlossene_aktie(wb, ws, row)

    return row


# ============================================================
# GESCHLOSSENE TRADES ARCHIV
# ============================================================

def _find_geschlossene_next_row(ws: Worksheet) -> int:
    """Nächste freie Zeile im Archiv (append ans Ende)."""
    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    return ws.max_row + 1


def _append_to_geschlossene(wb: Workbook, data: Dict[str, Any]):
    """Trägt einen geschlossenen Derivate-/ETF-Trade ins Archiv ein."""
    ws = wb[SHEET_GESCHLOSSEN]
    row = _find_geschlossene_next_row(ws)

    ws.cell(row=row, column=1).value = data['nr']
    ws.cell(row=row, column=2).value = data['kaufdatum']
    ws.cell(row=row, column=3).value = data['verkaufsdatum']
    ws.cell(row=row, column=4).value = data['instrument']
    ws.cell(row=row, column=5).value = data['isin']
    ws.cell(row=row, column=6).value = data['typ']
    ws.cell(row=row, column=7).value = data['richtung']

    c_h = ws.cell(row=row, column=8)
    c_h.value = data['kaufsumme']
    c_h.number_format = FMT_EURO

    c_i = ws.cell(row=row, column=9)
    c_i.value = data['erloes']
    c_i.number_format = FMT_EURO

    gv = data['gv']
    c_j = ws.cell(row=row, column=10)
    c_j.value = gv
    c_j.number_format = FMT_EURO
    if isinstance(gv, (int, float)):
        c_j.font = _gv_font(gv)

    c_k = ws.cell(row=row, column=11)
    c_k.value = data['gv_pct']
    c_k.number_format = FMT_PROZENT
    if isinstance(data['gv_pct'], (int, float)):
        c_k.font = _gv_font(data['gv_pct'])

    ws.cell(row=row, column=12).value = data['notizen']

    # Gebühren Spalten M + N (ab v3)
    gk = data.get('gebuehr_kauf')
    gv_fee = data.get('gebuehr_verkauf')
    if gk is not None:
        c_m = ws.cell(row=row, column=13)
        c_m.value = float(gk)
        c_m.number_format = FMT_EURO
    if gv_fee is not None:
        c_n = ws.cell(row=row, column=14)
        c_n.value = float(gv_fee)
        c_n.number_format = FMT_EURO


def _append_to_geschlossene_aktie(wb: Workbook, ws_av: Worksheet, av_row: int):
    """Trägt eine geschlossene Aktie ins Archiv ein (Typ='Aktie')."""
    ws = wb[SHEET_GESCHLOSSEN]
    row = _find_geschlossene_next_row(ws)

    nr = ws_av.cell(row=av_row, column=1).value
    kauf = ws_av.cell(row=av_row, column=2).value
    verk = ws_av.cell(row=av_row, column=3).value
    name = ws_av.cell(row=av_row, column=4).value
    isin = ws_av.cell(row=av_row, column=5).value
    einsatz = ws_av.cell(row=av_row, column=10).value
    erloes = ws_av.cell(row=av_row, column=11).value
    gv = ws_av.cell(row=av_row, column=12).value
    notizen = ws_av.cell(row=av_row, column=14).value
    gebuehr_kauf = ws_av.cell(row=av_row, column=15).value
    gebuehr_verkauf = ws_av.cell(row=av_row, column=16).value

    if isinstance(einsatz, (int, float)) and einsatz and isinstance(erloes, (int, float)):
        pct = (erloes - einsatz) / einsatz
    else:
        pct = None

    ws.cell(row=row, column=1).value = nr
    ws.cell(row=row, column=2).value = kauf
    ws.cell(row=row, column=3).value = verk
    ws.cell(row=row, column=4).value = name
    ws.cell(row=row, column=5).value = isin
    ws.cell(row=row, column=6).value = 'Aktie'
    ws.cell(row=row, column=7).value = 'Long'

    c_h = ws.cell(row=row, column=8)
    c_h.value = einsatz
    c_h.number_format = FMT_EURO

    c_i = ws.cell(row=row, column=9)
    c_i.value = erloes
    c_i.number_format = FMT_EURO

    c_j = ws.cell(row=row, column=10)
    c_j.value = gv
    c_j.number_format = FMT_EURO
    if isinstance(gv, (int, float)):
        c_j.font = _gv_font(gv)

    if pct is not None:
        c_k = ws.cell(row=row, column=11)
        c_k.value = pct
        c_k.number_format = FMT_PROZENT
        c_k.font = _gv_font(pct)

    ws.cell(row=row, column=12).value = notizen

    # Gebühren in M + N (v3)
    if gebuehr_kauf is not None:
        c_m = ws.cell(row=row, column=13)
        c_m.value = float(gebuehr_kauf)
        c_m.number_format = FMT_EURO
    if gebuehr_verkauf is not None:
        c_n = ws.cell(row=row, column=14)
        c_n.value = float(gebuehr_verkauf)
        c_n.number_format = FMT_EURO


# ============================================================
# WATCHLIST
# ============================================================

def add_watchlist(wb: Workbook, entry: Dict[str, Any]) -> int:
    """
    Fügt Eintrag zur Watchlist hinzu.
    Keys: **aktie**, **richtung** ('LONG'/'SHORT'), **trigger**, **these**, status (default '👀 beobachten'), datum (default heute)
    """
    ws = wb[SHEET_WATCHLIST]
    row = _find_geschlossene_next_row(ws) if False else 2
    # Finde erste leere Zeile ab Z2
    for r in range(2, ws.max_row + 2):
        if ws.cell(row=r, column=1).value is None:
            row = r
            break

    ws.cell(row=row, column=1).value = entry['aktie']
    ws.cell(row=row, column=2).value = entry['richtung']
    ws.cell(row=row, column=3).value = entry['trigger']
    ws.cell(row=row, column=4).value = entry['these']
    ws.cell(row=row, column=5).value = entry.get('status', '👀 beobachten')
    ws.cell(row=row, column=6).value = entry.get('datum', '')

    for col in range(1, 7):
        ws.cell(row=row, column=col).font = FONT_CALIBRI
    return row


def update_watchlist_status(wb: Workbook, aktie: str, status: str) -> bool:
    """Setzt neuen Status für Watchlist-Eintrag. Match per Substring."""
    ws = wb[SHEET_WATCHLIST]
    aktie_lower = aktie.lower()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and aktie_lower in val.lower():
            ws.cell(row=row, column=5).value = status
            return True
    return False


def remove_watchlist(wb: Workbook, aktie: str) -> bool:
    """Entfernt einen Watchlist-Eintrag (Match per Substring)."""
    ws = wb[SHEET_WATCHLIST]
    aktie_lower = aktie.lower()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and aktie_lower in val.lower():
            ws.delete_rows(row)
            return True
    return False


# --- Watchlist-Abgleich (für News-/Hidden-/Insider-Scans, Morgen-Briefing) ---

_WATCHLIST_STOPWORDS = frozenset({
    'ag', 'se', 'nv', 'inc', 'corp', 'gmbh', 'co', 'kg', 'vz', 'the',
    'und', 'and', 'ord', 'com', 'plc', 'ltd', 'spa', 'asa', 'sarl',
    'holding', 'holdings', 'group', 'aktien', 'aktie'
})


def _wl_normalize(s: str) -> str:
    """Normalisiert: lowercase, Umlaute → ae/oe/ue/ss, Sonderzeichen → Leerzeichen."""
    import re
    if not isinstance(s, str):
        return ''
    s = s.lower()
    s = s.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return s.strip()


def _wl_tokenize(s: str) -> set:
    """
    Tokenisiert auf signifikante Tokens (>=3 Zeichen, ohne Stopwords).
    Zusätzlich: Klammer-Kürzel wie '(EUZ)', '(AIXA)' werden als eigene Tokens behandelt.
    """
    tokens = set(_wl_normalize(s).split())
    return {t for t in tokens if len(t) >= 3 and t not in _WATCHLIST_STOPWORDS}


def list_watchlist(wb: Workbook) -> List[Dict[str, Any]]:
    """
    Gibt alle Watchlist-Einträge als Liste von Dicts zurück.

    Returns:
        Liste mit Keys: row, aktie, richtung, trigger, these, status, datum
    """
    ws = wb[SHEET_WATCHLIST]
    out: List[Dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        aktie = ws.cell(row=row, column=1).value
        if not aktie:
            continue
        out.append({
            'row': row,
            'aktie': aktie,
            'richtung': ws.cell(row=row, column=2).value or '',
            'trigger': ws.cell(row=row, column=3).value or '',
            'these': ws.cell(row=row, column=4).value or '',
            'status': ws.cell(row=row, column=5).value or '',
            'datum': ws.cell(row=row, column=6).value or '',
        })
    return out


def match_watchlist(wb: Workbook, query: str) -> Optional[Dict[str, Any]]:
    """
    Fuzzy-Match eines Kandidaten gegen die Watchlist.

    Matching-Strategie:
      Token-Match über signifikante Tokens (>=3 Zeichen, ohne Stopwords).
      Match wenn mind. 1 Token aus Query exakt einem Token im Aktien-Namen
      entspricht (fängt Kürzel wie 'EUZ', 'AIXA', 'GV6' ab, wenn sie im
      Journal als '(EUZ)', '(AIXA)' hinterlegt sind).

      Bewusst KEIN Substring-Match auf Zeichenebene — 'RWE' wäre sonst
      False-Positive-Match auf 'Drägerwerk' (erw**rwe**rk).

    Args:
        query: Aktienname, Kürzel, Ticker, WKN oder ISIN

    Returns:
        Dict mit Keys: row, aktie, richtung, trigger, these, status, datum
        oder None wenn kein Match.
        Bei mehreren Matches wird der erste (= oberste Watchlist-Zeile) zurückgegeben.

    Beispiele:
        match_watchlist(wb, 'AIXTRON')         → AIXTRON-Eintrag
        match_watchlist(wb, 'AIXA')            → AIXTRON-Eintrag (Kürzel)
        match_watchlist(wb, 'Eckert & Ziegler')→ EUZ-Eintrag
        match_watchlist(wb, 'EUZ')             → EUZ-Eintrag
        match_watchlist(wb, 'Deutsche Börse')  → Deutsche Börse AG
    """
    q_tokens = _wl_tokenize(query) if query else set()
    if not q_tokens:
        return None

    for e in list_watchlist(wb):
        entry_tokens = _wl_tokenize(e['aktie'])
        if q_tokens & entry_tokens:
            return e

    return None


# ============================================================
# ÜBERSICHT — TIMESTAMP + PORTFOLIO-BLOCK
# ============================================================

def update_timestamp(wb: Workbook, grund: str, datum: str = None):
    """
    Aktualisiert Stand-Zeile (A4) und PORTFOLIO-Überschrift (A16) mit aktuellem Datum.
    """
    from datetime import datetime
    if datum is None:
        datum = datetime.now().strftime('%d.%m.%Y')

    ws = wb[SHEET_UEBERSICHT]
    ws.cell(row=4, column=1).value = f"Stand: {datum} – {grund}"
    ws.cell(row=4, column=2).value = datum

    # Portfolio-Überschrift mit Datum
    ws.cell(row=16, column=1).value = f"PORTFOLIO — Offene Positionen (Stand {datum})"


def collect_open_positions(wb: Workbook) -> List[Dict[str, Any]]:
    """
    Sammelt alle OFFEN-Positionen aus SK + AV (nicht: TEIL-EXIT, GESCHLOSSEN, VERBUCHT).
    Für jede Position werden die Werte zurückgegeben, die in den Portfolio-Block gehören.
    """
    positions = []

    # SK: Spalte L = Status, bei OFFEN nehmen
    ws_sk = wb[SHEET_SK]
    rows_saldo = find_saldo_rows(ws_sk)
    for row in range(5, rows_saldo['saldo']):
        nr = ws_sk.cell(row=row, column=1).value
        status = ws_sk.cell(row=row, column=12).value
        if isinstance(nr, int) and status == 'OFFEN':
            positions.append({
                'nr': nr,
                'instrument_raw': ws_sk.cell(row=row, column=4).value,
                'typ': ws_sk.cell(row=row, column=6).value,
                'kaufsumme': ws_sk.cell(row=row, column=8).value,
                'notiz': ws_sk.cell(row=row, column=13).value or '',
                'source': 'SK',
            })

    # AV: Spalte M = Status
    ws_av = wb[SHEET_AV]
    for row in range(5, ws_av.max_row + 1):
        nr = ws_av.cell(row=row, column=1).value
        status = ws_av.cell(row=row, column=13).value
        if isinstance(nr, int) and status == 'OFFEN':
            positions.append({
                'nr': nr,
                'instrument_raw': ws_av.cell(row=row, column=4).value,
                'typ': 'Aktie',
                'kaufsumme': ws_av.cell(row=row, column=10).value,
                'notiz': ws_av.cell(row=row, column=14).value or '',
                'source': 'AV',
            })

    return positions


def update_portfolio_row(
    wb: Workbook,
    nr: int,
    buy_in: Optional[float] = None,
    tp1: Optional[float] = None,
    tp2: Optional[float] = None,
    sl_ori: Optional[float] = None,
    sl_empf: Optional[float] = None,
    zeitstopp: Optional[str] = None,
    wert: Optional[float] = None,
) -> bool:
    """
    Aktualisiert einzelne Felder (TP1/TP2/SL/Zeitstopp) in der Portfolio-Zeile der
    Übersicht für die Position mit gegebener Trade-Nr.
    Jedes None-Feld bleibt unverändert.
    """
    ws = wb[SHEET_UEBERSICHT]
    summe_row = find_summe_offen_row(ws)
    for row in range(PORTFOLIO_START_ROW, summe_row or ws.max_row):
        pos = ws.cell(row=row, column=1).value
        if isinstance(pos, str) and f"#{nr}" in pos:
            if wert is not None:
                ws.cell(row=row, column=2).value = float(wert)
                ws.cell(row=row, column=2).number_format = FMT_EURO
            if buy_in is not None:
                ws.cell(row=row, column=3).value = float(buy_in)
                ws.cell(row=row, column=3).number_format = FMT_EURO
            if tp1 is not None:
                ws.cell(row=row, column=4).value = float(tp1)
                ws.cell(row=row, column=4).number_format = FMT_EURO
            if tp2 is not None:
                ws.cell(row=row, column=5).value = float(tp2)
                ws.cell(row=row, column=5).number_format = FMT_EURO
            if sl_ori is not None:
                ws.cell(row=row, column=6).value = float(sl_ori)
                ws.cell(row=row, column=6).number_format = FMT_EURO
            if sl_empf is not None:
                ws.cell(row=row, column=7).value = float(sl_empf)
                ws.cell(row=row, column=7).number_format = FMT_EURO
            if zeitstopp is not None:
                ws.cell(row=row, column=8).value = zeitstopp
            return True
    return False


def remove_portfolio_row(wb: Workbook, nr: int) -> bool:
    """
    Entfernt die Portfolio-Zeile einer geschlossenen Position aus der Übersicht.
    Aktualisiert danach SUMME OFFEN.
    """
    ws = wb[SHEET_UEBERSICHT]
    summe_row = find_summe_offen_row(ws)
    if summe_row is None:
        return False

    for row in range(PORTFOLIO_START_ROW, summe_row):
        pos = ws.cell(row=row, column=1).value
        if isinstance(pos, str) and f"#{nr}" in pos:
            ws.delete_rows(row)
            _update_summe_offen(wb)
            return True
    return False


def _update_summe_offen(wb: Workbook):
    """Rechnet die SUMME OFFEN-Zeile neu aus Spalte B des Portfolio-Blocks."""
    ws = wb[SHEET_UEBERSICHT]
    summe_row = find_summe_offen_row(ws)
    if summe_row is None:
        return
    total = 0.0
    for row in range(PORTFOLIO_START_ROW, summe_row):
        val = ws.cell(row=row, column=2).value
        if isinstance(val, (int, float)):
            total += val
    c = ws.cell(row=summe_row, column=2)
    c.value = round(total, 2)
    c.number_format = FMT_EURO


def add_portfolio_row(
    wb: Workbook,
    nr: int,
    instrument: str,
    kurzname: str,
    wert: float,
    buy_in: Optional[float] = None,
    tp1: Optional[float] = None,
    tp2: Optional[float] = None,
    sl_ori: Optional[float] = None,
    sl_empf: Optional[float] = None,
    zeitstopp: Optional[str] = None,
    status: str = 'OFFEN',
    stk: Optional[int] = None,
) -> int:
    """
    Fügt eine neue Portfolio-Zeile in die Übersicht ein (direkt vor SUMME OFFEN).
    Zebra-Streifen (hellgrau auf ungerade Position im Block) wird beibehalten.
    """
    ws = wb[SHEET_UEBERSICHT]
    summe_row = find_summe_offen_row(ws)
    if summe_row is None:
        raise ValueError("SUMME OFFEN-Zeile nicht gefunden — Portfolio-Block defekt?")

    # Neue Zeile vor SUMME OFFEN einfügen
    new_row = summe_row
    ws.insert_rows(new_row)

    # Label zusammenbauen
    suffix = f", {stk} Stk" if stk is not None else ''
    label = f"{kurzname} (#{nr}{suffix})"

    c_a = ws.cell(row=new_row, column=1)
    c_a.value = label
    c_a.font = FONT_CALIBRI_BOLD

    c_b = ws.cell(row=new_row, column=2)
    c_b.value = float(wert)
    c_b.number_format = FMT_EURO

    def _num(col, v):
        c = ws.cell(row=new_row, column=col)
        if v is None:
            c.value = '—'
        else:
            c.value = float(v)
            c.number_format = FMT_EURO

    _num(3, buy_in)
    _num(4, tp1)
    _num(5, tp2)
    _num(6, sl_ori)
    _num(7, sl_empf)

    ws.cell(row=new_row, column=8).value = zeitstopp or '—'

    c_i = ws.cell(row=new_row, column=9)
    c_i.value = status
    c_i.font = FONT_CALIBRI_BOLD

    # Zebra-Streifen: zählt die Position im Block (1 = Kopf-Zeile nach Header = keine Zebra,
    # 2 = Zebra, 3 = keine, 4 = Zebra, …). Zwei Zeilen über Header (17) = Z18 = 1.
    position_in_block = new_row - PORTFOLIO_HEADER_ROW  # 1, 2, 3, …
    if position_in_block % 2 == 1:
        for col in range(1, 10):
            ws.cell(row=new_row, column=col).fill = FILL_ZEBRA

    _update_summe_offen(wb)
    return new_row


# ============================================================
# HIGH-LEVEL WRAPPER
# ============================================================

def close_trade_complete(
    wb: Workbook,
    nr: int,
    verkaufsdatum: str,
    erloes: float,
    lektion: Optional[str] = None,
    grund_timestamp: Optional[str] = None,
    kind: str = 'derivate',
    gebuehr_verkauf: Optional[float] = None,
) -> Dict[str, Any]:
    """
    Kompletter Exit-Workflow in einem Aufruf:
      1. Detail-Sheet Zeile schließen (SK oder AV)
      2. Saldo-Block neu berechnen (bei SK)
      3. Eintrag ins Archiv "Geschlossene Trades"
      4. Entsprechende Portfolio-Zeile aus Übersicht entfernen
      5. Übersicht-Timestamp aktualisieren

    kind: 'derivate' (SK Sheet) oder 'aktie' (AV Sheet)
    grund_timestamp: Kurzer Satz für A4 — z.B. 'HDD Short ausgestoppt'

    gebuehr_verkauf: Verkauf-Gebühr in € für die Transparenz-Spalte. Sollte
    bereits im übergebenen erloes abgezogen sein. Wenn None, bleibt die Zelle
    leer — am Jahresende sollte das aufgeräumt sein, sonst fehlt sie in der
    Gebühren-Summenzeile der Übersicht.
    """
    if kind == 'derivate':
        close_derivate_trade(wb, nr, verkaufsdatum, erloes, lektion=lektion,
                              archiv=True, gebuehr_verkauf=gebuehr_verkauf)
    elif kind == 'aktie':
        ws_av = wb[SHEET_AV]
        row = find_trade_row(ws_av, nr, status_filter='OFFEN')
        if row is None:
            raise ValueError(f"Keine OFFENE Aktie #{nr}")
        stueck = float(ws_av.cell(row=row, column=7).value)
        verkaufspreis_stk = float(erloes) / stueck if stueck else 0.0
        close_aktie(wb, nr, verkaufsdatum, verkaufspreis_stk, erloes=erloes,
                    lektion=lektion, archiv=True, gebuehr_verkauf=gebuehr_verkauf)
    else:
        raise ValueError(f"Unbekannter kind: {kind}. Erwarte 'derivate' oder 'aktie'.")

    remove_portfolio_row(wb, nr)

    if grund_timestamp:
        update_timestamp(wb, grund=grund_timestamp, datum=verkaufsdatum)


def add_trade_complete(
    wb: Workbook,
    trade: Dict[str, Any],
    portfolio_kurzname: str,
    portfolio_buy_in: Optional[float] = None,
    portfolio_tp1: Optional[float] = None,
    portfolio_tp2: Optional[float] = None,
    portfolio_sl_ori: Optional[float] = None,
    portfolio_sl_empf: Optional[float] = None,
    portfolio_zeitstopp: Optional[str] = None,
    portfolio_stk: Optional[int] = None,
    kind: str = 'derivate',
    grund_timestamp: Optional[str] = None,
) -> int:
    """
    Kompletter Entry-Workflow:
      1. Detail-Sheet Zeile anlegen (SK oder AV), OFFEN + gelb
      2. Portfolio-Zeile in Übersicht einfügen
      3. Übersicht-Timestamp aktualisieren

    Returns: Trade-Nr der neuen Position.
    """
    if kind == 'derivate':
        nr = add_derivate_trade(wb, trade)
    elif kind == 'aktie':
        nr = add_aktie(wb, trade)
    else:
        raise ValueError(f"Unbekannter kind: {kind}")

    add_portfolio_row(
        wb, nr=nr, instrument=trade.get('instrument', trade.get('aktie', '')),
        kurzname=portfolio_kurzname, wert=trade['kaufsumme'] if kind == 'derivate'
        else float(trade['stueck']) * float(trade['kaufpreis_stk']) + float(trade.get('gebuehr', 1.0)),
        buy_in=portfolio_buy_in, tp1=portfolio_tp1, tp2=portfolio_tp2,
        sl_ori=portfolio_sl_ori, sl_empf=portfolio_sl_empf,
        zeitstopp=portfolio_zeitstopp, status='OFFEN', stk=portfolio_stk,
    )

    if grund_timestamp:
        update_timestamp(wb, grund=grund_timestamp, datum=trade['kaufdatum'])
    return nr


# ============================================================
# NOTES — Handlungsbedarf, Milestones, offene Klärungen
# ============================================================
#
# Sheet "Notes" ersetzt die frühere externe JOURNAL_STATE.md.
# Alles, was nicht schon aus Übersicht/Watchlist/Geschlossene Trades
# ablesbar ist (TODOs, Milestones, offene ⚠️), lebt hier.
#
# Spaltenlayout:
#   A: ID (laufende Nummer)
#   B: Datum-erstellt (String dd.mm.yyyy)
#   C: Kategorie (TODO / ⚠️ / MILESTONE / INFO)
#   D: Notiz-Text
#   E: Status (OFFEN / ERLEDIGT)
#   F: Erledigt-am (String dd.mm.yyyy)
#
# Kategorien-Konvention:
#   TODO      — konkrete Handlung steht aus
#   ⚠️        — Klärungsbedarf / Verifikation (z.B. ISIN prüfen)
#   MILESTONE — chat-übergreifender Meilenstein (z.B. Trade #59 Review)
#   INFO      — kein Handlungsbedarf, aber merkwürdig

NOTES_HEADER = ['ID', 'Datum', 'Kategorie', 'Notiz', 'Status', 'Erledigt-am']
NOTES_KATEGORIEN = ('TODO', '⚠️', 'MILESTONE', 'INFO')


def ensure_notes_sheet(wb: Workbook) -> Worksheet:
    """
    Stellt sicher, dass das Notes-Sheet existiert. Legt es mit Header an,
    wenn nicht vorhanden. Idempotent — kann in jedem Workflow vorweg laufen.
    """
    if SHEET_NOTES in wb.sheetnames:
        return wb[SHEET_NOTES]

    ws = wb.create_sheet(SHEET_NOTES)
    for col, label in enumerate(NOTES_HEADER, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = label
        cell.font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_WEISS)
        cell.fill = FILL_HEADER_BLAU
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Spaltenbreiten sinnvoll
    widths = {'A': 6, 'B': 12, 'C': 12, 'D': 80, 'E': 11, 'F': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    return ws


def _next_note_id(ws: Worksheet) -> int:
    """Nächste freie ID in Notes-Sheet."""
    max_id = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, int) and val > max_id:
            max_id = val
    return max_id + 1


def _find_next_note_row(ws: Worksheet) -> int:
    """Nächste freie Zeile (ab 2)."""
    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    return ws.max_row + 1


def add_note(
    wb: Workbook,
    text: str,
    kategorie: str = 'TODO',
    datum: Optional[str] = None,
) -> int:
    """
    Fügt eine Notiz zum Notes-Sheet hinzu.

    Args:
        wb: Workbook
        text: Der Notiz-Text (frei)
        kategorie: 'TODO' | '⚠️' | 'MILESTONE' | 'INFO' (default 'TODO')
        datum: dd.mm.yyyy — default heute

    Returns:
        ID der neuen Notiz
    """
    from datetime import datetime
    if kategorie not in NOTES_KATEGORIEN:
        raise ValueError(f"Kategorie '{kategorie}' nicht erlaubt. Erlaubt: {NOTES_KATEGORIEN}")
    if datum is None:
        datum = datetime.now().strftime('%d.%m.%Y')

    ws = ensure_notes_sheet(wb)
    note_id = _next_note_id(ws)
    row = _find_next_note_row(ws)

    ws.cell(row=row, column=1).value = note_id
    ws.cell(row=row, column=2).value = datum
    ws.cell(row=row, column=3).value = kategorie
    ws.cell(row=row, column=4).value = text
    ws.cell(row=row, column=5).value = 'OFFEN'
    ws.cell(row=row, column=6).value = ''

    for col in range(1, 7):
        ws.cell(row=row, column=col).font = FONT_CALIBRI
    ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
    return note_id


def resolve_note(
    wb: Workbook,
    match: Any,
    datum: Optional[str] = None,
) -> bool:
    """
    Markiert Notiz als ERLEDIGT.

    Args:
        match: ID (int) ODER Substring des Notiz-Texts (case-insensitive)
        datum: dd.mm.yyyy — default heute

    Returns:
        True wenn gefunden und aktualisiert, sonst False
    """
    from datetime import datetime
    if datum is None:
        datum = datetime.now().strftime('%d.%m.%Y')
    if SHEET_NOTES not in wb.sheetnames:
        return False
    ws = wb[SHEET_NOTES]

    for row in range(2, ws.max_row + 1):
        nid = ws.cell(row=row, column=1).value
        txt = ws.cell(row=row, column=4).value
        status = ws.cell(row=row, column=5).value
        if status == 'ERLEDIGT':
            continue
        hit = False
        if isinstance(match, int) and nid == match:
            hit = True
        elif isinstance(match, str) and isinstance(txt, str) and match.lower() in txt.lower():
            hit = True
        if hit:
            ws.cell(row=row, column=5).value = 'ERLEDIGT'
            ws.cell(row=row, column=6).value = datum
            return True
    return False


def list_notes(
    wb: Workbook,
    nur_offen: bool = True,
    kategorie: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Gibt alle Notizen als Liste von Dicts zurück.

    Args:
        nur_offen: nur OFFEN-Notizen (default True)
        kategorie: Filter auf 'TODO' / '⚠️' / 'MILESTONE' / 'INFO' (default: alle)

    Returns:
        Liste von Dicts mit Keys: id, datum, kategorie, text, status, erledigt_am
    """
    if SHEET_NOTES not in wb.sheetnames:
        return []
    ws = wb[SHEET_NOTES]
    result = []
    for row in range(2, ws.max_row + 1):
        nid = ws.cell(row=row, column=1).value
        if nid is None:
            continue
        status = ws.cell(row=row, column=5).value
        kat = ws.cell(row=row, column=3).value
        if nur_offen and status != 'OFFEN':
            continue
        if kategorie and kat != kategorie:
            continue
        result.append({
            'id': nid,
            'datum': ws.cell(row=row, column=2).value,
            'kategorie': kat,
            'text': ws.cell(row=row, column=4).value,
            'status': status,
            'erledigt_am': ws.cell(row=row, column=6).value,
        })
    return result
